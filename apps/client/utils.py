from typing import Tuple, List

import openpyxl

from django.core.files.uploadedfile import InMemoryUploadedFile


def get_data_from_xls(file: InMemoryUploadedFile) -> list:
    wb = openpyxl.load_workbook(file)

    data = list()
    for row in wb['Лист1'].iter_rows():
        row_data = list()
        for cell in row:
            if cell.value:
                row_data.append(str(cell.value))
        data.append(row_data)
    print(data)
    return data


def get_data_from_client_xls(file : InMemoryUploadedFile) -> Tuple[List, List]:
    wb = openpyxl.load_workbook(file)

    client = list()
    organization = list()
    for row in wb['client'].iter_rows():
        row_data = list()
        for cell in row:
            if cell.value:
                row_data.append(str(cell.value))
        client.append(row_data)

    for row in wb['organization'].iter_rows():
        row_data = list()
        for cell in row:
            if cell.value:
                row_data.append(str(cell.value))
        organization.append(row_data)

    return client, organization
